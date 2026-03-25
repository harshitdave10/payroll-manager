from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from datetime import datetime, timedelta
from collections import defaultdict
import os, io, re, tempfile

app = Flask(__name__)

# Allow requests from GitHub Pages frontend (set FRONTEND_ORIGIN env var on Render)
frontend_origin = os.environ.get("FRONTEND_ORIGIN", "*")
CORS(app, origins=[frontend_origin])

VALID_USER = os.environ.get("PAYROLL_USER", "admin")
VALID_PASS = os.environ.get("PAYROLL_PASS", "payroll@2026")

# ── Styles ────────────────────────────────────────────────────────
header_font = Font(bold=True, name="Arial")
data_font   = Font(name="Arial", size=10)
center      = Alignment(horizontal="center", vertical="center")
thin        = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"),  bottom=Side(style="thin"))
header_fill = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
total_fill  = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
emp_fill    = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")


# ── Auth ──────────────────────────────────────────────────────────
@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json()
    if data.get("username") == VALID_USER and data.get("password") == VALID_PASS:
        return jsonify({"success": True, "message": "Login successful"})
    return jsonify({"success": False, "message": "Invalid credentials"}), 401


def normalize_text(value):
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def detect_excel_engine(filepath):
    with open(filepath, "rb") as f:
        signature = f.read(8)
    if signature.startswith(b"\xD0\xCF\x11\xE0"):
        return "xlrd"
    return "openpyxl"


def read_excel(filepath, **kwargs):
    return pd.read_excel(filepath, engine=detect_excel_engine(filepath), **kwargs)


def safe_sheet_name(name, used_names=None):
    cleaned = re.sub(r"[\[\]\*:/\\\?]", " ", normalize_text(name)) or "Sheet"
    cleaned = cleaned[:31]
    if used_names is None:
        return cleaned

    candidate = cleaned
    counter = 2
    while candidate in used_names:
        suffix = f"_{counter}"
        candidate = f"{cleaned[:31-len(suffix)]}{suffix}"
        counter += 1
    used_names.add(candidate)
    return candidate


def to_date_key(dt):
    return dt.strftime("%Y-%m-%d")


def parse_date_key(date_key):
    return datetime.strptime(date_key, "%Y-%m-%d")


def parse_report_period(df):
    pattern = re.compile(r"(\d{2}-[A-Za-z]{3}-\d{4})\s*-\s*(\d{2}-[A-Za-z]{3}-\d{4})")
    for i in range(min(len(df), 15)):
        for value in df.iloc[i].tolist():
            text = normalize_text(value)
            match = pattern.search(text)
            if match:
                return (
                    datetime.strptime(match.group(1), "%d-%b-%Y"),
                    datetime.strptime(match.group(2), "%d-%b-%Y"),
                )
    return (None, None)


def parse_header_date(header_value, year_hint):
    text = normalize_text(header_value)
    match = re.match(r"(\d{2}-[A-Za-z]{3})", text)
    if not match or not year_hint:
        return None
    try:
        return datetime.strptime(f"{match.group(1)}-{year_hint}", "%d-%b-%Y")
    except ValueError:
        return None


def format_cell_value(value):
    if pd.isna(value):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, timedelta):
        total_minutes = int(round(value.total_seconds() / 60))
        return f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"
    if isinstance(value, (int, float)) and 0 <= value < 1:
        total_minutes = int(round(float(value) * 24 * 60))
        return f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"

    text = normalize_text(value)
    if not text or text in {"-", "na", "NA", "nan", "NaT"}:
        return ""
    if text.count(":") == 2:
        return text[:5]
    return text


def parse_status_and_remark(value):
    text = format_cell_value(value)
    if not text:
        return "", ""
    if "|" in text:
        parts = [part.strip() for part in text.split("|") if part.strip()]
        if parts:
            return parts[0], " | ".join(parts[1:])
    return text, ""


def normalize_description(value):
    text = normalize_text(value).lower()
    if text.startswith("status"):
        return "status"
    if text in {"in", "in time", "intime", "check in", "check-in"}:
        return "in"
    if text.startswith("out"):
        return "out"
    if "worked" in text:
        return "worked"
    if "break" in text or "lunch" in text:
        return "lunch"
    return text


def find_column_index(headers, predicate):
    for idx, header in enumerate(headers):
        if predicate(normalize_text(header).lower()):
            return idx
    return None


def first_nonempty_value(rows, col_idx):
    if col_idx is None:
        return ""
    for _, row in rows.iterrows():
        value = format_cell_value(row.iloc[col_idx])
        if value:
            return value
    return ""


def parse_input_sheet(filepath):
    df = read_excel(filepath, header=None)
    report_start, _ = parse_report_period(df)

    header_row = None
    for i in range(min(len(df), 25)):
        headers = [normalize_text(v).lower() for v in df.iloc[i].tolist()]
        if any("base site" in h for h in headers) and "name" in headers and "designation" in headers and "description" in headers:
            header_row = i
            break

    if header_row is None:
        raise ValueError("Could not find the Truein header row in the input file")

    headers = list(df.iloc[header_row].tolist())
    site_idx = find_column_index(headers, lambda h: "base site" in h)
    emp_id_idx = find_column_index(headers, lambda h: h in {"employee-id", "employee id"})
    name_idx = find_column_index(headers, lambda h: h == "name")
    designation_idx = find_column_index(headers, lambda h: h == "designation")
    description_idx = find_column_index(headers, lambda h: h == "description")
    overtime_idx = find_column_index(headers, lambda h: h in {"overtime hrs", "overtime hours", "ot hours", "ot hrs"})
    payable_days_idx = find_column_index(headers, lambda h: h == "payable days")

    if None in (site_idx, name_idx, designation_idx, description_idx):
        raise ValueError("Input file is missing required columns")

    date_columns = []
    for idx in range(description_idx + 1, len(headers)):
        dt = parse_header_date(headers[idx], report_start.year if report_start else None)
        if dt:
            date_columns.append((idx, dt))

    if not date_columns:
        raise ValueError("Could not find any date columns in the input file")

    rows = df.iloc[header_row + 1:].copy()
    rows = rows.dropna(how="all")
    for idx in [site_idx, emp_id_idx, name_idx, designation_idx]:
        if idx is not None:
            rows.iloc[:, idx] = rows.iloc[:, idx].ffill()

    site_employees = defaultdict(list)
    site_dates = defaultdict(set)

    group_cols = [site_idx, name_idx, designation_idx]
    if emp_id_idx is not None:
        group_cols.insert(1, emp_id_idx)

    grouped = rows.groupby(group_cols, sort=False, dropna=False)

    for group_key, group_rows in grouped:
        site_name = normalize_text(group_rows.iloc[0, site_idx])
        employee_name = normalize_text(group_rows.iloc[0, name_idx])
        designation = normalize_text(group_rows.iloc[0, designation_idx])
        employee_id = normalize_text(group_rows.iloc[0, emp_id_idx]) if emp_id_idx is not None else ""
        if not site_name or not employee_name:
            continue

        row_map = {}
        for _, row in group_rows.iterrows():
            description_key = normalize_description(row.iloc[description_idx])
            if description_key:
                row_map[description_key] = row

        if not any(key in row_map for key in ("status", "in", "out", "worked", "lunch")):
            continue

        emp_data = {}
        for col_idx, dt in date_columns:
            status, remark = parse_status_and_remark(row_map.get("status", pd.Series()).iloc[col_idx] if "status" in row_map else "")
            in_time = format_cell_value(row_map.get("in", pd.Series()).iloc[col_idx] if "in" in row_map else "")
            out_time = format_cell_value(row_map.get("out", pd.Series()).iloc[col_idx] if "out" in row_map else "")
            worked = format_cell_value(row_map.get("worked", pd.Series()).iloc[col_idx] if "worked" in row_map else "")
            lunch = format_cell_value(row_map.get("lunch", pd.Series()).iloc[col_idx] if "lunch" in row_map else "")

            date_key = to_date_key(dt)
            site_dates[site_name].add(date_key)
            emp_data[date_key] = {
                "status": status,
                "in": in_time,
                "out": out_time,
                "worked": worked,
                "ot": "",
                "lunch": lunch,
                "remark": remark,
            }

        site_employees[site_name].append({
            "site": site_name,
            "employee_id": employee_id,
            "name": employee_name,
            "designation": designation,
            "overtime_total": first_nonempty_value(group_rows, overtime_idx),
            "payable_days": first_nonempty_value(group_rows, payable_days_idx),
            "data": emp_data,
        })

    return site_employees, {
        site: sorted(date_keys, key=parse_date_key) for site, date_keys in site_dates.items()
    }


def coerce_output_date(value):
    if pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value
    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime()
    if isinstance(value, (int, float)) and 30000 <= float(value) <= 60000:
        return datetime(1899, 12, 30) + timedelta(days=float(value))
    text = normalize_text(value)
    for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def parse_existing_output(filepath):
    sheets = pd.read_excel(filepath, sheet_name=None, header=None, engine="openpyxl")
    site_employees = defaultdict(list)
    site_dates = defaultdict(set)

    for site_name, df in sheets.items():
        i = 0
        while i < len(df) - 1:
            row = df.iloc[i]
            next_row = df.iloc[i + 1]
            col0 = normalize_text(row[0]) if len(row) > 0 else ""
            next0 = normalize_text(next_row[0]).upper() if len(next_row) > 0 else ""

            if col0 and next0 == "DATE":
                designation = normalize_text(row[5]) if len(row) > 5 else ""
                employee = {
                    "site": site_name,
                    "employee_id": "",
                    "name": col0,
                    "designation": designation,
                    "overtime_total": "",
                    "payable_days": "",
                    "data": {},
                }

                j = i + 2
                while j < len(df):
                    drow = df.iloc[j]
                    label = normalize_text(drow[0]) if len(drow) > 0 else ""
                    if label == "TOTAL DAYS":
                        employee["payable_days"] = format_cell_value(drow[3]) if len(drow) > 3 else ""
                        if j + 1 < len(df):
                            employee["overtime_total"] = format_cell_value(df.iloc[j + 1][3]) if len(df.iloc[j + 1]) > 3 else ""
                        break

                    date_val = coerce_output_date(drow[0] if len(drow) > 0 else None)
                    if date_val:
                        date_key = to_date_key(date_val)
                        site_dates[site_name].add(date_key)
                        employee["data"][date_key] = {
                            "status": format_cell_value(drow[1]) if len(drow) > 1 else "",
                            "in": format_cell_value(drow[2]) if len(drow) > 2 else "",
                            "out": format_cell_value(drow[3]) if len(drow) > 3 else "",
                            "worked": format_cell_value(drow[4]) if len(drow) > 4 else "",
                            "ot": format_cell_value(drow[5]) if len(drow) > 5 else "",
                            "lunch": format_cell_value(drow[6]) if len(drow) > 6 else "",
                            "remark": format_cell_value(drow[7]) if len(drow) > 7 else "",
                        }
                    j += 1

                site_employees[site_name].append(employee)
                i = j + 1
            else:
                i += 1

    return site_employees, {
        site: sorted(date_keys, key=parse_date_key) for site, date_keys in site_dates.items()
    }


def write_employee_block(ws, start_row, emp, date_keys):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=5)
    ws.merge_cells(start_row=start_row, start_column=6, end_row=start_row, end_column=8)

    name_cell = ws.cell(start_row, 1, emp["name"])
    name_cell.font = Font(bold=True, name="Arial")
    name_cell.fill = emp_fill
    name_cell.alignment = center

    designation_cell = ws.cell(start_row, 6, emp["designation"])
    designation_cell.font = Font(bold=True, name="Arial")
    designation_cell.fill = emp_fill
    designation_cell.alignment = center

    for col in range(1, 9):
        ws.cell(start_row, col).border = thin

    header_row = start_row + 1
    headers = ["DATE", "STATUS", "IN TIME", "OUT TIME", "DUTY HOURS", "OT", "LUNCH", "REMARK"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin

    data_start = header_row + 1
    for idx, date_key in enumerate(date_keys):
        row_idx = data_start + idx
        record = emp["data"].get(date_key, {})
        date_cell = ws.cell(row_idx, 1, parse_date_key(date_key))
        date_cell.number_format = "DD-MMM-YYYY"
        date_cell.font = data_font
        date_cell.border = thin
        date_cell.alignment = center

        values = [
            record.get("status", ""),
            record.get("in", ""),
            record.get("out", ""),
            record.get("worked", ""),
            record.get("ot", ""),
            record.get("lunch", ""),
            record.get("remark", ""),
        ]
        for col_offset, value in enumerate(values, start=2):
            cell = ws.cell(row_idx, col_offset, value)
            cell.font = data_font
            cell.border = thin
            cell.alignment = center

    total_row = data_start + len(date_keys)
    for label, value, row_idx in [
        ("TOTAL DAYS", emp.get("payable_days", ""), total_row),
        ("OVER TIME", emp.get("overtime_total", ""), total_row + 1),
        ("LUNCH", "0", total_row + 2),
    ]:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        label_cell = ws.cell(row_idx, 1, label)
        value_cell = ws.cell(row_idx, 4, value)
        for cell in (label_cell, value_cell):
            cell.font = Font(bold=True, name="Arial")
            cell.fill = total_fill
            cell.border = thin
            cell.alignment = center

    summary_row = total_row + 3
    for col, value in [(1, "OP. BALANCE LEAVE"), (3, "LEAVE TEKAN THIS MONTH"), (6, "BALANCE LEAVE")]:
        cell = ws.cell(summary_row, col, value)
        cell.font = Font(bold=True, name="Arial")
        cell.fill = total_fill
        cell.border = thin
        cell.alignment = center

    return summary_row + 2


def build_workbook(site_employees, site_dates):
    wb = Workbook()
    wb.remove(wb.active)
    used_sheet_names = set()

    for site_name, employees in site_employees.items():
        ws = wb.create_sheet(title=safe_sheet_name(site_name, used_sheet_names))
        for i, width in enumerate([14, 12, 10, 10, 12, 10, 10, 18], start=1):
            ws.column_dimensions[chr(64 + i)].width = width

        current_row = 1
        ordered_dates = site_dates.get(site_name) or sorted({
            date_key for employee in employees for date_key in employee["data"].keys()
        }, key=parse_date_key)

        for employee in employees:
            current_row = write_employee_block(ws, current_row, employee, ordered_dates)

    if not wb.sheetnames:
        ws = wb.create_sheet(title="Payroll")
        ws["A1"] = "No employee data found"

    return wb


def employee_key(employee):
    return (
        normalize_text(employee.get("name")).lower(),
        normalize_text(employee.get("designation")).lower(),
    )


# ── New Month ─────────────────────────────────────────────────────
@app.route("/api/new-month", methods=["POST"])
def new_month():
    if "input_file" not in request.files:
        return jsonify({"error": "No input file provided"}), 400

    f   = request.files["input_file"]
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    f.save(tmp.name); tmp.close()

    try:
        site_employees, site_dates = parse_input_sheet(tmp.name)
        if not site_employees:
            return jsonify({"error": "No employee data found in input file"}), 400

        wb  = build_workbook(site_employees, site_dates)
        out = io.BytesIO()
        wb.save(out); out.seek(0)
        return send_file(out, as_attachment=True,
                         download_name="final_payroll.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        os.unlink(tmp.name)


# ── Existing Month ────────────────────────────────────────────────
@app.route("/api/existing-month", methods=["POST"])
def existing_month():
    if "input_file" not in request.files or "output_file" not in request.files:
        return jsonify({"error": "Both input and output files are required"}), 400

    tmp_in  = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_out = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    request.files["input_file"].save(tmp_in.name);   tmp_in.close()
    request.files["output_file"].save(tmp_out.name); tmp_out.close()

    try:
        new_sites, new_site_dates = parse_input_sheet(tmp_in.name)
        if not new_sites:
            return jsonify({"error": "No employee data found in input file"}), 400

        existing_sites, existing_site_dates = parse_existing_output(tmp_out.name)

        merged_sites = defaultdict(list)
        merged_dates = defaultdict(set)
        all_sites = set(existing_sites.keys()) | set(new_sites.keys())

        for site_name in all_sites:
            existing_employees = list(existing_sites.get(site_name, []))
            new_employees = list(new_sites.get(site_name, []))
            existing_map = {employee_key(emp): emp for emp in existing_employees}

            for new_emp in new_employees:
                key = employee_key(new_emp)
                if key in existing_map:
                    existing_emp = existing_map[key]
                    for date_key, day_data in new_emp["data"].items():
                        if date_key not in existing_emp["data"]:
                            existing_emp["data"][date_key] = day_data
                    if new_emp.get("overtime_total"):
                        existing_emp["overtime_total"] = new_emp["overtime_total"]
                    if new_emp.get("payable_days"):
                        existing_emp["payable_days"] = new_emp["payable_days"]
                    if new_emp.get("designation"):
                        existing_emp["designation"] = new_emp["designation"]
                else:
                    existing_employees.append(new_emp)

            merged_sites[site_name] = existing_employees
            merged_dates[site_name].update(existing_site_dates.get(site_name, []))
            merged_dates[site_name].update(new_site_dates.get(site_name, []))

        wb  = build_workbook(
            merged_sites,
            {site: sorted(date_keys, key=parse_date_key) for site, date_keys in merged_dates.items()},
        )
        out = io.BytesIO()
        wb.save(out); out.seek(0)
        return send_file(out, as_attachment=True,
                         download_name="final_payroll.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        os.unlink(tmp_in.name)
        os.unlink(tmp_out.name)


if __name__ == "__main__":
    print(f"Starting Payroll Server — user: {VALID_USER}")
    app.run(debug=False, port=5000)
